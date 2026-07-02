from openpyxl import load_workbook

HEADER_MAP = {
    "invoice_number": ["Invoice No", "Invoice Number", "Inv No"],
    "invoice_date": ["Invoice Date", "Date"],
    "supplier": ["Supplier", "Vendor", "Company"],
    "sap_number": ["SAP Number", "SAP No", "SAP"],
    "quantity": ["Quantity", "Qty"]
}


def append_to_excel(excel_path, invoice):

    wb = load_workbook(excel_path)
    ws = wb.active

    # Read header row
    headers = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value

        if value is None:
            continue

        value = str(value).strip().lower()

        for key, possible_names in HEADER_MAP.items():
            if value in [name.lower() for name in possible_names]:
                headers[key] = col

    next_row = ws.max_row + 1

    # Fill matching columns
    for key, col in headers.items():
        ws.cell(row=next_row, column=col).value = invoice.get(key, "")

    wb.save(excel_path)